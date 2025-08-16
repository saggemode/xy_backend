from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.contrib import messages
from django.utils import timezone
from .models import Wallet, Transaction, BankTransfer, BillPayment, VirtualCard, Bank, VATCharge, TransferChargeControl, CBNLevy, TransactionCharge, NightGuardSettings, LargeTransactionShieldSettings, LocationGuardSettings, MerchantSettlementAccount
import json
from django.utils.safestring import mark_safe
from django.urls import reverse
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.template.loader import render_to_string
# from weasyprint import HTML
from django.utils import timezone
from django import forms
from django.urls import path
from django.template.response import TemplateResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count, Q
from .models import Transaction, Wallet
from .models import (
    Wallet, Transaction, BankTransfer, BillPayment, VirtualCard, Bank, 
    TransferFailure,
    StaffRole, StaffProfile, TransactionApproval, CustomerEscalation, StaffActivity,
    XySaveAccount, XySaveTransaction, XySaveGoal, XySaveInvestment, XySaveSettings,
    SpendAndSaveAccount, SpendAndSaveTransaction, SpendAndSaveSettings,
    TargetSaving, TargetSavingDeposit, TargetSavingCategory, TargetSavingFrequency, TargetSavingWithdrawal,
    FixedSavingsAccount, FixedSavingsTransaction, FixedSavingsSettings,
    FixedSavingsSource, FixedSavingsPurpose
)


from djmoney.models.fields import MoneyField
from djmoney.money import Money
from .interest_services import InterestRateCalculator, InterestAccrualService, InterestReportService

@admin.register(Wallet)
class WalletAdmin(admin.ModelAdmin):
    list_display = ('user', 'account_number', 'alternative_account_number', 'balance', 'currency', 'created_at', 'balance_status')
    list_filter = ('currency', 'created_at')
    search_fields = ('user__username', 'user__email', 'account_number', 'alternative_account_number')
    readonly_fields = ('created_at', 'updated_at', 'interest_info', 'interest_breakdown')
    ordering = ('-created_at',)
    list_per_page = 25
    
    fieldsets = (
        ('User Information', {
            'fields': ('user',)
        }),
        ('Account Details', {
            'fields': ('account_number', 'alternative_account_number')
        }),
        ('Financial Information', {
            'fields': ('balance', 'currency')
        }),
        ('Interest Information', {
            'fields': ('interest_info', 'interest_breakdown'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def balance_status(self, obj):
        """Display balance status with color coding."""
        if obj.balance > Money(0, obj.balance.currency):
            return format_html(
                '<span style="color: #28a745; font-weight: bold;">₦{}</span>',
                f"{obj.balance.amount:,.2f}"
            )
        else:
            return format_html(
                '<span style="color: #dc3545; font-weight: bold;">₦{}</span>',
                f"{obj.balance.amount:,.2f}"
            )
    balance_status.short_description = 'Balance'
    
    def interest_info(self, obj):
        """Display interest information for the wallet."""
        if obj.balance.amount <= 0:
            return format_html('<span style="color: #6c757d;">No interest (zero balance)</span>')
        
        # Calculate annual interest
        annual_result = InterestRateCalculator.calculate_interest_breakdown(obj.balance, 365)
        effective_rate = float(annual_result['effective_rate'] * 100)
        
        return format_html(
            '<div>'
            '<strong>Annual Interest:</strong> ₦{annual_interest}<br>'
            '<strong>Effective Rate:</strong> {effective_rate:.2f}%<br>'
            '<strong>Monthly Interest:</strong> ₦{monthly_interest}'
            '</div>',
            annual_interest=f"{annual_result['total_interest'].amount:,.2f}",
            effective_rate=effective_rate,
            monthly_interest=f"{InterestRateCalculator.calculate_interest_for_balance(obj.balance, 30).amount:,.2f}"
        )
    interest_info.short_description = 'Interest Info'
    
    def interest_breakdown(self, obj):
        """Display detailed interest breakdown."""
        if obj.balance.amount <= 0:
            return "No interest calculation for zero balance"
        
        result = InterestRateCalculator.calculate_interest_breakdown(obj.balance, 365)
        
        breakdown_html = '<div style="background: #f8f9fa; padding: 10px; border-radius: 5px;">'
        breakdown_html += '<h4>Annual Interest Breakdown:</h4>'
        
        for tier in result['breakdown']:
            breakdown_html += f'<p><strong>Tier {tier["tier"]}:</strong> {tier["balance_in_tier"]} at {tier["rate"]*100}% p.a. = {tier["interest"]}</p>'
        
        breakdown_html += f'<hr><p><strong>Total Annual Interest:</strong> {result["total_interest"]}</p>'
        breakdown_html += f'<p><strong>Effective Annual Rate:</strong> {float(result["effective_rate"] * 100):.2f}%</p>'
        breakdown_html += '</div>'
        
        return mark_safe(breakdown_html)


@admin.register(NightGuardSettings)
class NightGuardSettingsAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'enabled', 'start_time', 'end_time', 'primary_method',
        'fallback_method', 'applies_to', 'face_registered_at'
    )
    list_filter = ('enabled', 'primary_method', 'fallback_method', 'applies_to')
    search_fields = ('user__username', 'user__email')
    readonly_fields = ('face_registered_at',)
    fieldsets = (
        ('User', {'fields': ('user',)}),
        ('Window', {'fields': ('enabled', 'start_time', 'end_time', 'applies_to')}),
        ('Verification', {'fields': ('primary_method', 'fallback_method')}),
        ('Face Enrollment', {'fields': ('face_template_alg', 'face_template_hash', 'face_registered_at')}),
        ('Timestamps', {'fields': ('created_at', 'updated_at')}),
    )
    readonly_fields = ('created_at', 'updated_at', 'face_registered_at')


@admin.register(LargeTransactionShieldSettings)
class LargeTransactionShieldSettingsAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'enabled', 'per_transaction_limit', 'daily_limit', 'monthly_limit', 'face_registered_at'
    )
    list_filter = ('enabled',)
    search_fields = ('user__username', 'user__email')
    fieldsets = (
        ('User', {'fields': ('user',)}),
        ('Limits', {'fields': ('enabled', 'per_transaction_limit', 'daily_limit', 'monthly_limit')}),
        ('Face Enrollment', {'fields': ('face_template_alg', 'face_template_hash', 'face_registered_at')}),
        ('Timestamps', {'fields': ('created_at', 'updated_at')}),
    )
    readonly_fields = ('created_at', 'updated_at', 'face_registered_at')


@admin.register(LocationGuardSettings)
class LocationGuardSettingsAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'enabled', 'face_registered_at'
    )
    list_filter = ('enabled',)
    search_fields = ('user__username', 'user__email')
    fieldsets = (
        ('User', {'fields': ('user',)}),
        ('Allowed States', {'fields': ('enabled', 'allowed_states')}),
        ('Face Enrollment', {'fields': ('face_template_alg', 'face_template_hash', 'face_registered_at')}),
        ('Timestamps', {'fields': ('created_at', 'updated_at')}),
    )
    readonly_fields = ('created_at', 'updated_at', 'face_registered_at')


@admin.register(MerchantSettlementAccount)
class MerchantSettlementAccountAdmin(admin.ModelAdmin):
    list_display = ('merchant_id', 'bank_code', 'account_number', 'account_name', 'is_verified', 'preferred_schedule', 'updated_at')
    list_filter = ('is_verified', 'preferred_schedule')
    search_fields = ('merchant_id', 'account_number', 'account_name')
    readonly_fields = ('created_at', 'updated_at', 'is_verified', 'verification_method')
    
    actions = ['add_balance', 'deduct_balance', 'export_wallet_data', 'calculate_interest', 'apply_interest']
    
    def add_balance(self, request, queryset):
        """Add balance to selected wallets."""
        # This would typically open a form to specify amount
        self.message_user(request, f'Add balance action for {queryset.count()} wallets.')
    add_balance.short_description = "Add balance to selected wallets"
    
    def deduct_balance(self, request, queryset):
        """Deduct balance from selected wallets."""
        self.message_user(request, f'Deduct balance action for {queryset.count()} wallets.')
    deduct_balance.short_description = "Deduct balance from selected wallets"
    
    def export_wallet_data(self, request, queryset):
        """Export wallet data."""
        self.message_user(request, f'Exporting {queryset.count()} wallet records...')
    export_wallet_data.short_description = "Export wallet data"
    
    def calculate_interest(self, request, queryset):
        """Calculate interest for selected wallets."""
        total_interest = Money(0, 'NGN')
        for wallet in queryset:
            if wallet.balance.amount > 0:
                annual_interest = InterestRateCalculator.calculate_interest_for_balance(wallet.balance, 365)
                total_interest += annual_interest
        
        self.message_user(
            request, 
            f'Calculated annual interest for {queryset.count()} wallets. Total: {total_interest}'
        )
    calculate_interest.short_description = "Calculate interest for selected wallets"
    
    def apply_interest(self, request, queryset):
        """Apply monthly interest to selected wallets."""
        applied_count = 0
        total_applied = Money(0, 'NGN')
        
        for wallet in queryset:
            if wallet.balance.amount > 0:
                try:
                    transaction = InterestAccrualService.process_monthly_interest(wallet)
                    if transaction:
                        applied_count += 1
                        total_applied += transaction.amount
                except Exception as e:
                    self.message_user(request, f'Error applying interest to {wallet.account_number}: {str(e)}', level=messages.ERROR)
        
        if applied_count > 0:
            self.message_user(
                request, 
                f'Applied interest to {applied_count} wallets. Total applied: {total_applied}'
            )
        else:
            self.message_user(request, 'No interest was applied to any wallets.')
    apply_interest.short_description = "Apply monthly interest to selected wallets"

class HasReversalListFilter(admin.SimpleListFilter):
    title = 'Has Reversal'
    parameter_name = 'has_reversal'
    def lookups(self, request, model_admin):
        return (
            ('yes', 'Has Reversal'),
            ('no', 'No Reversal'),
        )
    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(reversals__isnull=False).distinct()
        if self.value() == 'no':
            return queryset.filter(reversals__isnull=True)
        return queryset

class LargeTransactionListFilter(admin.SimpleListFilter):
    title = 'Large Transaction'
    parameter_name = 'large_tx'
    def lookups(self, request, model_admin):
        return (
            ('yes', 'Large (>=₦100,000)'),
            ('no', 'Not Large'),
        )
    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(amount__amount__gte=100000)
        if self.value() == 'no':
            return queryset.filter(amount__amount__lt=100000)
        return queryset

class SuspiciousActivityListFilter(admin.SimpleListFilter):
    title = 'Suspicious Activity'
    parameter_name = 'suspicious'
    def lookups(self, request, model_admin):
        return (
            ('yes', 'Flagged as Suspicious'),
            ('no', 'Not Flagged'),
        )
    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(metadata__suspicious=True)
        if self.value() == 'no':
            return queryset.exclude(metadata__suspicious=True)
        return queryset

class ReversalInline(admin.TabularInline):
    model = Transaction
    fk_name = 'parent'
    extra = 0
    fields = ('reference', 'amount', 'type', 'channel', 'status', 'timestamp', 'currency')
    readonly_fields = fields
    show_change_link = True

class PDFStatementForm(forms.Form):
    date_from = forms.DateField(label='From', required=True, widget=forms.DateInput(attrs={'type': 'date'}))
    date_to = forms.DateField(label='To', required=True, widget=forms.DateInput(attrs={'type': 'date'}))

@admin.register(Transaction)
class TransactionAdmin(admin.ModelAdmin):
    list_display = (
        'reference', 'wallet', 'amount_display', 'type', 'channel', 'status',
        'timestamp', 'balance_after', 'currency', 'parent_summary', 'related_object_link', 'pretty_metadata', 'reversal_summary'
    )
    readonly_fields = ('balance_after', 'related_object_link', 'pretty_metadata', 'parent_summary', 'reversal_summary')
    search_fields = ('reference', 'wallet__user__username', 'description')
    list_filter = ('type', 'channel', 'status', 'currency', HasReversalListFilter, LargeTransactionListFilter, SuspiciousActivityListFilter)
    ordering = ('-timestamp',)
    list_per_page = 50
    inlines = [ReversalInline]

    fieldsets = (
        ('Transaction Details', {
            'fields': ('reference', 'wallet', 'amount', 'type', 'channel')
        }),
        ('Status Information', {
            'fields': ('status', 'description')
        }),
    )
    
    def amount_display(self, obj):
        """Display amount with color coding."""
        try:
            amount = obj.amount.amount if hasattr(obj.amount, 'amount') else float(obj.amount)
            if obj.type == 'credit':
                return format_html(
                    '<span style="color: #28a745; font-weight: bold;">+₦{}</span>',
                    f"{amount:,.2f}"
                )
            else:
                return format_html(
                    '<span style="color: #dc3545; font-weight: bold;">-₦{}</span>',
                    f"{amount:,.2f}"
                )
        except (AttributeError, TypeError, ValueError):
            # Fallback if amount is not a Money object or can't be formatted
            return str(obj.amount)
    amount_display.short_description = 'Amount'
    
    actions = ['mark_as_success', 'mark_as_failed', 'export_transactions', 'export_transactions_csv', 'export_transactions_xlsx', 'export_pdf_statement',
        'bulk_approve', 'bulk_review', 'bulk_flag_suspicious', 'bulk_reverse'
    ]
    
    def mark_as_success(self, request, queryset):
        """Mark transactions as successful."""
        updated = queryset.update(status='success')
        self.message_user(request, f'Marked {updated} transactions as successful.')
    mark_as_success.short_description = "Mark as successful"
    
    def mark_as_failed(self, request, queryset):
        """Mark transactions as failed."""
        updated = queryset.update(status='failed')
        self.message_user(request, f'Marked {updated} transactions as failed.')
    mark_as_failed.short_description = "Mark as failed"
    
    def export_transactions(self, request, queryset):
        """Export transaction data."""
        self.message_user(request, f'Exporting {queryset.count()} transaction records...')
    export_transactions.short_description = "Export transaction data"

    def export_transactions_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
        writer = csv.writer(response)
        writer.writerow(['Reference', 'Wallet', 'Amount', 'Type', 'Channel', 'Status', 'Timestamp', 'Balance After', 'Currency', 'Parent', 'Related Object', 'Metadata'])
        for obj in queryset:
            writer.writerow([
                obj.reference,
                str(obj.wallet),
                obj.amount,
                obj.type,
                obj.channel,
                obj.status,
                obj.timestamp,
                obj.balance_after,
                obj.currency,
                obj.parent.reference if obj.parent else '',
                str(obj.related_object) if obj.related_object else '',
                json.dumps(obj.metadata) if obj.metadata else '',
            ])
        return response
    export_transactions_csv.short_description = "Export Selected Transactions as CSV"

    def export_transactions_xlsx(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        headers = ['Reference', 'Wallet', 'Amount', 'Type', 'Channel', 'Status', 'Timestamp', 'Balance After', 'Currency', 'Parent', 'Related Object', 'Metadata']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for obj in queryset:
            ws.append([
                obj.reference,
                str(obj.wallet),
                obj.amount,
                obj.type,
                obj.channel,
                obj.status,
                obj.timestamp,
                obj.balance_after,
                obj.currency,
                obj.parent.reference if obj.parent else '',
                str(obj.related_object) if obj.related_object else '',
                json.dumps(obj.metadata, indent=2) if obj.metadata else '',
            ])
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
        wb.save(response)
        return response
    export_transactions_xlsx.short_description = "Export Selected Transactions as Excel (XLSX)"

    def export_pdf_statement(self, request, queryset):
        if 'apply' in request.POST:
            form = PDFStatementForm(request.POST)
            if form.is_valid():
                date_from = form.cleaned_data['date_from']
                date_to = form.cleaned_data['date_to']
                txs = queryset.filter(timestamp__date__gte=date_from, timestamp__date__lte=date_to).order_by('timestamp')
                if not txs.exists():
                    self.message_user(request, 'No transactions in selected range.', level=messages.WARNING)
                    return
                wallet = txs.first().wallet
                user = wallet.user
                opening_balance = txs.first().balance_after - txs.first().amount if txs.first().balance_after is not None else 0
                closing_balance = txs.last().balance_after if txs.last().balance_after is not None else 0
                total_credits = sum(t.amount for t in txs if t.type == 'credit')
                total_debits = sum(t.amount for t in txs if t.type == 'debit')
                html = render_to_string('bank/statement_pdf.html', {
                    'user': user,
                    'wallet': wallet,
                    'transactions': txs,
                    'date_from': date_from,
                    'date_to': date_to,
                    'opening_balance': opening_balance,
                    'closing_balance': closing_balance,
                    'total_credits': total_credits,
                    'total_debits': total_debits,
                    'now': timezone.now(),
                })
                # pdf_file = HTML(string=html).write_pdf() # Commented out to disable PDF generation
                # response = HttpResponse(pdf_file, content_type='application/pdf')
                # response['Content-Disposition'] = f'attachment; filename=statement_{wallet.account_number}_{date_from}_{date_to}.pdf'
                # return response
                self.message_user(request, 'PDF statement generation is currently disabled.', level=messages.WARNING)
                return
        else:
            form = PDFStatementForm()
        return admin.helpers.render_action_form(request, form, 'Export PDF Statement for Date Range')
    export_pdf_statement.short_description = 'Export PDF Statement for Date Range'

    def related_object_link(self, obj):
        if obj.related_object:
            ct = obj.content_type
            url = reverse(f'admin:{ct.app_label}_{ct.model}_change', args=[obj.object_id])
            return mark_safe(f'<a href="{url}">{obj.related_object}</a>')
        return "-"
    related_object_link.short_description = 'Related Object'

    def pretty_metadata(self, obj):
        if obj.metadata:
            return mark_safe(f'<pre>{json.dumps(obj.metadata, indent=2)}</pre>')
        return "-"
    pretty_metadata.short_description = 'Metadata'

    def parent_summary(self, obj):
        if obj.parent:
            return f"{obj.parent.reference} ({obj.parent.amount} {obj.parent.currency})"
        return "-"
    parent_summary.short_description = 'Parent Transaction'

    def reversal_summary(self, obj):
        reversals = obj.reversals.all()
        if reversals:
            return mark_safe('<br>'.join([f"{r.reference} ({r.amount} {r.currency})" for r in reversals]))
        return "-"
    reversal_summary.short_description = 'Reversals'

    def bulk_approve(self, request, queryset):
        updated = queryset.update(status='success')
        self.message_user(request, f'Approved {updated} transactions.')
    bulk_approve.short_description = 'Approve selected transactions'
    def bulk_review(self, request, queryset):
        for tx in queryset:
            if not tx.metadata:
                tx.metadata = {}
            tx.metadata['reviewed'] = True
            tx.save(update_fields=['metadata'])
        self.message_user(request, f'Marked {queryset.count()} transactions as reviewed.')
    bulk_review.short_description = 'Mark selected as reviewed'
    def bulk_flag_suspicious(self, request, queryset):
        for tx in queryset:
            if not tx.metadata:
                tx.metadata = {}
            tx.metadata['suspicious'] = True
            tx.save(update_fields=['metadata'])
        self.message_user(request, f'Flagged {queryset.count()} transactions as suspicious.')
    bulk_flag_suspicious.short_description = 'Flag selected as suspicious'
    def bulk_reverse(self, request, queryset):
        count = 0
        for tx in queryset:
            if not tx.parent:  # Only reverse original transactions
                tx.wallet.balance += tx.amount  # Refund
                tx.wallet.save()
                tx.status = 'success'
                tx.save(update_fields=['status'])
                from .models import Transaction
                Transaction.objects.create(
                    wallet=tx.wallet,
                    reference=f"REV-{tx.reference}",
                    amount=tx.amount,
                    type='credit' if tx.type == 'debit' else 'debit',
                    channel=tx.channel,
                    description=f"Reversal of {tx.reference}",
                    status='success',
                    currency=tx.currency,
                    parent=tx
                )
                count += 1
        self.message_user(request, f'Reversed {count} transactions.')
    bulk_reverse.short_description = 'Reverse selected transactions'
    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        # Advanced search: metadata fields, amount, date
        if search_term.startswith('meta:'):
            key, _, value = search_term[5:].partition('=')
            if key and value:
                queryset |= self.model.objects.filter(metadata__contains={key: value})
        if search_term.startswith('amount:'):
            _, _, value = search_term.partition(':')
            try:
                amount = float(value)
                queryset |= self.model.objects.filter(amount=amount)
            except:
                pass
        if search_term.startswith('date:'):
            _, _, value = search_term.partition(':')
            queryset |= self.model.objects.filter(timestamp__date=value)
        return queryset, use_distinct

@admin.register(BankTransfer)
class BankTransferAdmin(admin.ModelAdmin):
    list_display = (
        'reference', 'user', 'bank_name', 'account_number', 'amount', 'fee', 'vat', 'levy', 'transfer_type', 'description', 'nibss_reference', 'status_badge', 'created_at'
    )
    list_filter = ('status', 'created_at', 'bank_name', 'transfer_type')
    search_fields = ('reference', 'user__username', 'bank_name', 'account_number', 'description', 'nibss_reference')
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)
    list_per_page = 25

    fieldsets = (
        ('Transfer Details', {
            'fields': ('reference', 'user', 'bank_name', 'account_number', 'amount', 'fee', 'vat', 'levy', 'transfer_type', 'description', 'nibss_reference')
        }),
        ('Status', {
            'fields': ('status',)
        }),
        ('Timestamps', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )

    actions = ['mark_as_completed', 'mark_as_failed', 'export_transfers_csv', 'export_transfers_xlsx']

    def status_badge(self, obj):
        color = {
            'completed': '#28a745',
            'pending': '#ffc107',
            'failed': '#dc3545'
        }.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background: {}; color: #fff; padding: 2px 8px; border-radius: 6px;">{}</span>',
            color, obj.get_status_display()
        )
    status_badge.short_description = 'Status'

    def mark_as_completed(self, request, queryset):
        updated = queryset.update(status='completed')
        self.message_user(request, f'Marked {updated} transfers as completed.')
    mark_as_completed.short_description = "Mark as completed"

    def mark_as_failed(self, request, queryset):
        updated = queryset.update(status='failed')
        self.message_user(request, f'Marked {updated} transfers as failed.')
    mark_as_failed.short_description = "Mark as failed"

    def export_transfers_csv(self, request, queryset):
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="bank_transfers.csv"'
        writer = csv.writer(response)
        writer.writerow(['Reference', 'User', 'Bank Name', 'Account Number', 'Amount', 'Fee', 'VAT', 'Levy', 'Status', 'Created At'])
        for obj in queryset:
            writer.writerow([
                obj.reference,
                str(obj.user),
                obj.bank_name,
                obj.account_number,
                obj.amount,
                obj.fee,
                obj.vat,
                obj.levy,
                obj.get_status_display(),
                obj.created_at,
            ])
        return response
    export_transfers_csv.short_description = "Export Selected Transfers as CSV"

    def export_transfers_xlsx(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bank Transfers'
        headers = ['Reference', 'User', 'Bank Name', 'Account Number', 'Amount', 'Fee', 'VAT', 'Levy', 'Status', 'Created At']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for obj in queryset:
            ws.append([
                obj.reference,
                str(obj.user),
                obj.bank_name,
                obj.account_number,
                obj.amount,
                obj.fee,
                obj.vat,
                obj.levy,
                obj.get_status_display(),
                obj.created_at,
            ])
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=bank_transfers.xlsx'
        wb.save(response)
        return response
    export_transfers_xlsx.short_description = "Export Selected Transfers as Excel (XLSX)"

@admin.register(BillPayment)
class BillPaymentAdmin(admin.ModelAdmin):
    list_display = ('reference', 'user', 'service_type', 'account_or_meter', 'amount', 'status', 'timestamp')
    list_filter = ('status', 'service_type', 'timestamp')
    search_fields = ('reference', 'user__username', 'service_type', 'account_or_meter')
    readonly_fields = ('timestamp',)
    ordering = ('-timestamp',)
    
    fieldsets = (
        ('Payment Details', {
            'fields': ('reference', 'user', 'service_type', 'account_or_meter', 'amount')
        }),
        ('Status', {
            'fields': ('status',)
        }),
        ('Timestamps', {
            'fields': ('timestamp',),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['mark_as_success', 'mark_as_failed']

@admin.register(VirtualCard)
class VirtualCardAdmin(admin.ModelAdmin):
    list_display = ('user', 'card_number_masked', 'expiry', 'provider', 'status', 'issued_at')
    list_filter = ('status', 'provider', 'issued_at')
    search_fields = ('user__username', 'card_number', 'provider')
    readonly_fields = ('issued_at',)
    ordering = ('-issued_at',)
    
    def card_number_masked(self, obj):
        """Display masked card number."""
        return f"**** **** **** {obj.card_number[-4:]}"
    card_number_masked.short_description = 'Card Number'
    
    fieldsets = (
        ('Card Information', {
            'fields': ('user', 'card_number', 'expiry', 'cvv', 'provider')
        }),
        ('Status', {
            'fields': ('status',)
        }),
        ('Timestamps', {
            'fields': ('issued_at',),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['activate_cards', 'block_cards']

@admin.register(Bank)
class BankAdmin(admin.ModelAdmin):
    list_display = ('name', 'code', 'ussd', 'logo_preview', 'slug')
    list_filter = ('code',)
    search_fields = ('name', 'code', 'ussd')
    ordering = ('name',)
    readonly_fields = ('logo_preview',)
    
    fieldsets = (
        ('Bank Information', {
            'fields': ('name', 'code', 'slug')
        }),
        ('Additional Details', {
            'fields': ('ussd', 'logo', 'logo_preview'),
            'classes': ('collapse',)
        }),
    )
    
    def logo_preview(self, obj):
        """Display logo preview if available."""
        if obj.logo:
            return format_html(
                '<img src="{}" alt="{}" style="max-width: 50px; max-height: 30px; border-radius: 4px;" />',
                obj.logo, obj.name
            )
        return "No logo"
    logo_preview.short_description = 'Logo'

@admin.register(VATCharge)
class VATChargeAdmin(admin.ModelAdmin):
    list_display = ('rate', 'active', 'updated_at')
    list_filter = ('active',)
    ordering = ('-updated_at',)

@admin.register(CBNLevy)
class CBNLevyAdmin(admin.ModelAdmin):
    list_display = ('name', 'rate', 'transaction_type', 'is_active', 'effective_from')
    search_fields = ('name', 'regulation_reference')
    list_filter = ('is_active', 'transaction_type', 'effective_from')
    ordering = ('-effective_from',)
    
@admin.register(TransferChargeControl)
class TransferChargeControlAdmin(admin.ModelAdmin):
    list_display = ('levy_active', 'vat_active', 'fee_active', 'updated_at')
    ordering = ('-updated_at',)

@admin.register(TransactionCharge)
class TransactionChargeAdmin(admin.ModelAdmin):
    list_display = ('transfer', 'transfer_fee', 'vat_amount', 'levy_amount', 'charge_status', 'created_at')
    search_fields = ('transfer__reference',)
    list_filter = ('charge_status', 'created_at')
    ordering = ('-created_at',)

@admin.register(TransferFailure)
class TransferFailureAdmin(admin.ModelAdmin):
    list_display = (
        'transfer_reference', 'error_code', 'error_category', 'failure_reason', 
        'user_email', 'transfer_amount', 'failed_at', 'is_resolved'
    )
    list_filter = ('error_category', 'is_resolved', 'failed_at', 'error_code')
    search_fields = (
        'transfer__reference', 'failure_reason', 'user_id', 
        'recipient_account', 'technical_details'
    )
    readonly_fields = ('failed_at', 'created_at', 'updated_at')
    ordering = ('-failed_at',)
    list_per_page = 25
    
    fieldsets = (
        ('Transfer Information', {
            'fields': ('transfer', 'transfer_amount', 'recipient_account', 'recipient_bank_code')
        }),
        ('Failure Details', {
            'fields': ('error_code', 'error_category', 'failure_reason', 'technical_details', 'stack_trace')
        }),
        ('Context Information', {
            'fields': ('user_id', 'ip_address', 'user_agent', 'device_fingerprint'),
            'classes': ('collapse',)
        }),
        ('Resolution', {
            'fields': ('is_resolved', 'resolution_notes', 'resolved_by', 'resolved_at')
        }),
        ('Retry Information', {
            'fields': ('retry_count', 'last_retry_at', 'max_retries'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('failed_at', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['mark_resolved', 'increment_retry', 'export_failures']
    
    def transfer_reference(self, obj):
        """Display transfer reference with link."""
        if obj.transfer:
            return format_html(
                '<a href="{}">{}</a>',
                reverse('admin:bank_banktransfer_change', args=[obj.transfer.id]),
                obj.transfer.reference
            )
        return "N/A"
    transfer_reference.short_description = 'Transfer Reference'
    
    def user_email(self, obj):
        """Display user email from transfer."""
        if obj.transfer and obj.transfer.user:
            return obj.transfer.user.email
        return "N/A"
    user_email.short_description = 'User Email'
    
    def mark_resolved(self, request, queryset):
        """Mark selected failures as resolved."""
        updated = queryset.update(
            is_resolved=True,
            resolved_by=request.user,
            resolved_at=timezone.now()
        )
        self.message_user(request, f'{updated} failure(s) marked as resolved.')
    mark_resolved.short_description = "Mark as resolved"
    
    def increment_retry(self, request, queryset):
        """Increment retry count for selected failures."""
        for failure in queryset:
            failure.increment_retry()
        self.message_user(request, f'Retry count incremented for {queryset.count()} failure(s).')
    increment_retry.short_description = "Increment retry count"
    
    def export_failures(self, request, queryset):
        """Export failure data to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transfer_failures.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Transfer Reference', 'Error Code', 'Error Category', 'Failure Reason',
            'User Email', 'Transfer Amount', 'Recipient Account', 'Failed At', 'Is Resolved'
        ])
        
        for failure in queryset:
            writer.writerow([
                failure.transfer.reference if failure.transfer else 'N/A',
                failure.error_code,
                failure.error_category,
                failure.failure_reason,
                failure.transfer.user.email if failure.transfer and failure.transfer.user else 'N/A',
                failure.transfer_amount,
                failure.recipient_account,
                failure.failed_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Yes' if failure.is_resolved else 'No'
            ])
        
        return response
    export_failures.short_description = "Export to CSV"

# Skipping TransferChargeControl admin registration (model not present)

@admin.register(StaffRole)
class StaffRoleAdmin(admin.ModelAdmin):
    list_display = ('name', 'level', 'max_transaction_approval', 'permissions_summary', 'staff_count')
    list_filter = ('level', 'can_approve_kyc', 'can_manage_staff', 'can_view_reports')
    search_fields = ('name', 'description')
    ordering = ('level',)
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'level', 'description')
        }),
        ('Transaction Permissions', {
            'fields': ('max_transaction_approval', 'can_override_transactions')
        }),
        ('Administrative Permissions', {
            'fields': ('can_approve_kyc', 'can_manage_staff', 'can_view_reports', 'can_handle_escalations')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ('created_at', 'updated_at')
    
    def permissions_summary(self, obj):
        """Display a summary of permissions."""
        perms = []
        if obj.can_approve_kyc:
            perms.append('KYC')
        if obj.can_manage_staff:
            perms.append('Staff')
        if obj.can_view_reports:
            perms.append('Reports')
        if obj.can_override_transactions:
            perms.append('Override')
        if obj.can_handle_escalations:
            perms.append('Escalations')
        return ', '.join(perms) if perms else 'None'
    permissions_summary.short_description = 'Permissions'
    
    def staff_count(self, obj):
        """Count of staff members with this role."""
        return obj.staff_members.count()
    staff_count.short_description = 'Staff Count'


@admin.register(StaffProfile)
class StaffProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'role', 'employee_id', 'branch', 'department', 'supervisor', 'is_active', 'hire_date')
    list_filter = ('role', 'branch', 'department', 'is_active', 'hire_date')
    search_fields = ('user__username', 'user__email', 'user__first_name', 'user__last_name', 'employee_id')
    ordering = ('role__level', 'user__username')
    
    fieldsets = (
        ('User Information', {
            'fields': ('user', 'employee_id')
        }),
        ('Role & Branch', {
            'fields': ('role', 'branch', 'department')
        }),
        ('Supervision', {
            'fields': ('supervisor',)
        }),
        ('Status', {
            'fields': ('is_active', 'hire_date', 'last_review_date')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ('created_at', 'updated_at')
    
    actions = ['activate_staff', 'deactivate_staff', 'export_staff_data']
    
    def activate_staff(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'Activated {updated} staff members.')
    activate_staff.short_description = "Activate selected staff"
    
    def deactivate_staff(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'Deactivated {updated} staff members.')
    deactivate_staff.short_description = "Deactivate selected staff"
    
    def export_staff_data(self, request, queryset):
        self.message_user(request, f'Exporting {queryset.count()} staff records...')
    export_staff_data.short_description = "Export staff data"


@admin.register(TransactionApproval)
class TransactionApprovalAdmin(admin.ModelAdmin):
    list_display = ('transaction', 'requested_by', 'status', 'approved_by', 'created_at', 'escalated_to')
    list_filter = ('status', 'created_at', 'requested_by__role')
    search_fields = ('transaction__reference', 'requested_by__user__username', 'reason')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Transaction Information', {
            'fields': ('transaction', 'requested_by')
        }),
        ('Approval Details', {
            'fields': ('status', 'approved_by', 'reason')
        }),
        ('Escalation', {
            'fields': ('escalation_reason', 'escalated_to')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ('created_at', 'updated_at')
    
    actions = ['approve_transactions', 'reject_transactions', 'escalate_transactions']
    
    def approve_transactions(self, request, queryset):
        # This would typically require additional logic to check permissions
        updated = queryset.filter(status='pending').update(status='approved')
        self.message_user(request, f'Approved {updated} transactions.')
    approve_transactions.short_description = "Approve selected transactions"
    
    def reject_transactions(self, request, queryset):
        updated = queryset.filter(status='pending').update(status='rejected')
        self.message_user(request, f'Rejected {updated} transactions.')
    reject_transactions.short_description = "Reject selected transactions"
    
    def escalate_transactions(self, request, queryset):
        updated = queryset.filter(status='pending').update(status='escalated')
        self.message_user(request, f'Escalated {updated} transactions.')
    escalate_transactions.short_description = "Escalate selected transactions"


@admin.register(CustomerEscalation)
class CustomerEscalationAdmin(admin.ModelAdmin):
    list_display = ('subject', 'customer', 'priority', 'status', 'created_by', 'assigned_to', 'created_at')
    list_filter = ('priority', 'status', 'created_at', 'created_by__role')
    search_fields = ('subject', 'customer__username', 'description')
    ordering = ('-priority', '-created_at')
    
    fieldsets = (
        ('Customer Information', {
            'fields': ('customer', 'created_by')
        }),
        ('Escalation Details', {
            'fields': ('subject', 'description', 'priority', 'status')
        }),
        ('Assignment', {
            'fields': ('assigned_to',)
        }),
        ('Resolution', {
            'fields': ('resolution', 'resolved_at')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ('created_at', 'updated_at')
    
    actions = ['assign_to_self', 'mark_in_progress', 'mark_resolved', 'mark_closed']
    
    def assign_to_self(self, request, queryset):
        # Get the current user's staff profile
        try:
            staff_profile = StaffProfile.objects.get(user=request.user)
            updated = queryset.filter(status='open').update(assigned_to=staff_profile)
            self.message_user(request, f'Assigned {updated} escalations to yourself.')
        except StaffProfile.DoesNotExist:
            self.message_user(request, 'You do not have a staff profile.', level=messages.ERROR)
    assign_to_self.short_description = "Assign to self"
    
    def mark_in_progress(self, request, queryset):
        updated = queryset.filter(status='open').update(status='in_progress')
        self.message_user(request, f'Marked {updated} escalations as in progress.')
    mark_in_progress.short_description = "Mark as in progress"
    
    def mark_resolved(self, request, queryset):
        from django.utils import timezone
        updated = queryset.filter(status='in_progress').update(
            status='resolved', 
            resolved_at=timezone.now()
        )
        self.message_user(request, f'Marked {updated} escalations as resolved.')
    mark_resolved.short_description = "Mark as resolved"
    
    def mark_closed(self, request, queryset):
        updated = queryset.filter(status='resolved').update(status='closed')
        self.message_user(request, f'Marked {updated} escalations as closed.')
    mark_closed.short_description = "Mark as closed"


@admin.register(StaffActivity)
class StaffActivityAdmin(admin.ModelAdmin):
    list_display = ('staff', 'activity_type', 'description', 'timestamp')
    list_filter = ('activity_type', 'timestamp', 'staff__role')
    search_fields = ('staff__user__username', 'description')
    ordering = ('-timestamp',)
    
    fieldsets = (
        ('Activity Information', {
            'fields': ('staff', 'activity_type', 'description')
        }),
        ('Related Object', {
            'fields': ('related_object_type', 'related_object_id'),
            'classes': ('collapse',)
        }),
        ('Timestamp', {
            'fields': ('timestamp',),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ('timestamp',)
    
    def has_add_permission(self, request):
        """Staff activities should be created programmatically, not manually."""
        return False

@staff_member_required
def admin_dashboard(request):
    # Aggregate data for charts
    from django.utils import timezone
    import datetime
    today = timezone.now().date()
    last_30 = [today - datetime.timedelta(days=i) for i in range(29, -1, -1)]
    txs = Transaction.objects.filter(timestamp__date__gte=last_30[0])
    # Volume per day
    volume_data = []
    for day in last_30:
        count = txs.filter(timestamp__date=day).count()
        volume_data.append({'date': day.strftime('%Y-%m-%d'), 'count': count})
    # Revenue (sum of credits)
    revenue_data = []
    for day in last_30:
        revenue = txs.filter(timestamp__date=day, type='credit').aggregate(total=Sum('amount'))['total'] or 0
        revenue_data.append({'date': day.strftime('%Y-%m-%d'), 'revenue': float(revenue)})
    # Reversals per day
    reversal_data = []
    for day in last_30:
        reversals = txs.filter(timestamp__date=day, parent__isnull=False).count()
        reversal_data.append({'date': day.strftime('%Y-%m-%d'), 'count': reversals})
    # Top users by volume
    top_users = (
        Wallet.objects.values('user__username')
        .annotate(total=Sum('transaction__amount'))
        .order_by('-total')[:5]
    )
    context = {
        'volume_data': volume_data,
        'revenue_data': revenue_data,
        'reversal_data': reversal_data,
        'top_users': top_users,
    }
    return TemplateResponse(request, 'bank/admin_dashboard.html', context)

@staff_member_required
def interest_calculator_admin(request):
    """Admin view for interest rate calculator."""
    context = {
        'title': 'Interest Rate Calculator',
        'opts': Wallet._meta,
    }
    
    if request.method == 'POST':
        try:
            balance_amount = float(request.POST.get('balance', 0))
            days = int(request.POST.get('days', 365))
            currency = request.POST.get('currency', 'NGN')
            
            if balance_amount > 0:
                balance = Money(balance_amount, currency)
                result = InterestRateCalculator.calculate_interest_breakdown(balance, days)
                
                # Calculate additional periods
                result['monthly_interest'] = InterestRateCalculator.calculate_interest_for_balance(balance, 30).amount
                result['weekly_interest'] = InterestRateCalculator.calculate_interest_for_balance(balance, 7).amount
                result['daily_interest'] = InterestRateCalculator.calculate_interest_for_balance(balance, 1).amount
                
                context.update({
                    'calculation_result': result,
                    'input_balance': balance,
                    'input_days': days,
                    'input_currency': currency,
                })
            else:
                context['error'] = 'Balance must be greater than 0'
                
        except (ValueError, TypeError) as e:
            context['error'] = f'Invalid input: {str(e)}'
        except Exception as e:
            context['error'] = f'Calculation error: {str(e)}'
    
    # Get interest rates info
    context['rates_info'] = InterestReportService.get_interest_rates_info()
    
    return TemplateResponse(request, 'admin/bank/interest_calculator.html', context)

@staff_member_required
def staff_dashboard(request):
    """Staff management dashboard with role-based views and performance metrics."""
    from django.utils import timezone
    import datetime
    
    # Get current user's staff profile
    try:
        current_staff = StaffProfile.objects.get(user=request.user)
    except StaffProfile.DoesNotExist:
        return HttpResponse("Staff profile not found", status=404)
    
    today = timezone.now().date()
    last_30 = [today - datetime.timedelta(days=i) for i in range(29, -1, -1)]
    
    # Staff statistics
    total_staff = StaffProfile.objects.filter(is_active=True).count()
    staff_by_role = StaffRole.objects.annotate(
        staff_count=Count('staff_members', filter=Q(staff_members__is_active=True))
    )
    
    # Recent activities
    recent_activities = StaffActivity.objects.select_related('staff', 'staff__role').order_by('-timestamp')[:10]
    
    # Pending approvals
    pending_approvals = TransactionApproval.objects.filter(
        status='pending'
    ).select_related('transaction', 'requested_by', 'requested_by__role')[:10]
    
    # Open escalations
    open_escalations = CustomerEscalation.objects.filter(
        status__in=['open', 'in_progress']
    ).select_related('customer', 'created_by', 'assigned_to')[:10]
    
    # Staff performance (last 30 days)
    staff_performance = []
    for staff in StaffProfile.objects.filter(is_active=True)[:10]:
        activities_count = StaffActivity.objects.filter(
            staff=staff,
            timestamp__date__gte=last_30[0]
        ).count()
        
        approvals_count = TransactionApproval.objects.filter(
            approved_by=staff,
            status='approved',
            created_at__date__gte=last_30[0]
        ).count()
        
        escalations_resolved = CustomerEscalation.objects.filter(
            assigned_to=staff,
            status='resolved',
            resolved_at__date__gte=last_30[0]
        ).count()
        
        staff_performance.append({
            'staff': staff,
            'activities': activities_count,
            'approvals': approvals_count,
            'escalations_resolved': escalations_resolved,
        })
    
    # Role-based data access
    if current_staff.can_manage_staff():
        # Managers can see all staff data
        subordinates = current_staff.get_subordinates()
        staff_under_management = StaffProfile.objects.filter(supervisor=current_staff)
    else:
        # Regular staff can only see their own data
        subordinates = StaffProfile.objects.none()
        staff_under_management = StaffProfile.objects.none()
    
    context = {
        'current_staff': current_staff,
        'total_staff': total_staff,
        'staff_by_role': staff_by_role,
        'recent_activities': recent_activities,
        'pending_approvals': pending_approvals,
        'open_escalations': open_escalations,
        'staff_performance': staff_performance,
        'subordinates': subordinates,
        'staff_under_management': staff_under_management,
        'can_manage_staff': current_staff.can_manage_staff(),
        'can_view_reports': current_staff.role.can_view_reports,
    }
    
    return TemplateResponse(request, 'bank/staff_dashboard.html', context)


# XySave Admin Models

@admin.register(XySaveAccount)
class XySaveAccountAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'account_number', 'balance', 'total_interest_earned', 
        'daily_interest_rate', 'annual_interest_rate', 'auto_save_enabled', 
        'is_active', 'created_at'
    )
    list_filter = ('is_active', 'auto_save_enabled', 'created_at')
    search_fields = ('user__username', 'user__email', 'account_number')
    readonly_fields = ('created_at', 'updated_at', 'annual_interest_rate', 'daily_interest_calc')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Account Information', {
            'fields': ('user', 'account_number')
        }),
        ('Financial Information', {
            'fields': ('balance', 'total_interest_earned', 'daily_interest_rate')
        }),
        ('Auto-Save Settings', {
            'fields': ('auto_save_enabled', 'auto_save_percentage', 'auto_save_min_amount')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
        ('Interest Information', {
            'fields': ('annual_interest_rate', 'daily_interest_calc'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def annual_interest_rate(self, obj):
        """Display annual interest rate"""
        return f"{obj.get_annual_interest_rate():.2f}%"
    annual_interest_rate.short_description = 'Annual Rate'
    
    def daily_interest_calc(self, obj):
        """Display calculated daily interest"""
        daily_interest = obj.calculate_daily_interest()
        return f"₦{daily_interest.amount:,.2f}"
    daily_interest_calc.short_description = 'Daily Interest'
    
    actions = ['enable_auto_save', 'disable_auto_save', 'calculate_interest']
    
    def enable_auto_save(self, request, queryset):
        """Enable auto-save for selected accounts"""
        updated = queryset.update(auto_save_enabled=True)
        self.message_user(request, f"Auto-save enabled for {updated} accounts.")
    enable_auto_save.short_description = "Enable Auto-Save"
    
    def disable_auto_save(self, request, queryset):
        """Disable auto-save for selected accounts"""
        updated = queryset.update(auto_save_enabled=False)
        self.message_user(request, f"Auto-save disabled for {updated} accounts.")
    disable_auto_save.short_description = "Disable Auto-Save"
    
    def calculate_interest(self, request, queryset):
        """Calculate interest for selected accounts"""
        from .xysave_services import XySaveInterestService
        XySaveInterestService.calculate_daily_interest_for_all_accounts()
        self.message_user(request, "Daily interest calculated for all active accounts.")
    calculate_interest.short_description = "Calculate Daily Interest"


@admin.register(XySaveTransaction)
class XySaveTransactionAdmin(admin.ModelAdmin):
    list_display = (
        'reference', 'xysave_account', 'transaction_type', 'amount', 
        'balance_before', 'balance_after', 'created_at'
    )
    list_filter = ('transaction_type', 'created_at')
    search_fields = ('reference', 'xysave_account__user__username', 'description')
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Transaction Details', {
            'fields': ('reference', 'xysave_account', 'transaction_type', 'amount')
        }),
        ('Balance Information', {
            'fields': ('balance_before', 'balance_after')
        }),
        ('Additional Information', {
            'fields': ('description', 'metadata')
        }),
        ('Timestamp', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )


@admin.register(XySaveGoal)
class XySaveGoalAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'user', 'target_amount', 'current_amount', 
        'is_active', 'created_at'
    )
    list_filter = ('is_active', 'created_at')
    search_fields = ('name', 'user__username')
    readonly_fields = ('created_at', 'updated_at', 'progress_percentage', 'is_completed')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Goal Information', {
            'fields': ('name', 'user', 'target_amount', 'target_date')
        }),
        ('Progress', {
            'fields': ('current_amount', 'progress_percentage', 'is_completed')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def progress_percentage(self, obj):
        """Display progress percentage with color coding"""
        percentage = obj.get_progress_percentage()
        if percentage >= 100:
            color = '#28a745'  # Green
        elif percentage >= 75:
            color = '#17a2b8'  # Blue
        elif percentage >= 50:
            color = '#ffc107'  # Yellow
        else:
            color = '#dc3545'  # Red
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}%</span>',
            color, f"{percentage:.1f}"
        )
    progress_percentage.short_description = 'Progress'
    
    def is_completed(self, obj):
        """Display completion status"""
        return obj.is_completed()
    is_completed.boolean = True
    is_completed.short_description = 'Completed'


@admin.register(XySaveInvestment)
class XySaveInvestmentAdmin(admin.ModelAdmin):
    list_display = (
        'investment_type', 'xysave_account', 'amount_invested', 'current_value',
        'return_percentage', 'expected_return_rate', 'is_active', 'created_at'
    )
    list_filter = ('investment_type', 'is_active', 'created_at')
    search_fields = ('xysave_account__user__username', 'investment_type')
    readonly_fields = ('created_at', 'updated_at', 'return_percentage')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Investment Details', {
            'fields': ('xysave_account', 'investment_type', 'amount_invested', 'expected_return_rate')
        }),
        ('Current Status', {
            'fields': ('current_value', 'return_percentage', 'maturity_date')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def return_percentage(self, obj):
        """Display return percentage with color coding"""
        percentage = obj.get_return_percentage()
        if percentage > 0:
            color = '#28a745'  # Green
        elif percentage < 0:
            color = '#dc3545'  # Red
        else:
            color = '#6c757d'  # Gray
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}%</span>',
            color, f"{percentage:.2f}"
        )
    return_percentage.short_description = 'Return %'


@admin.register(XySaveSettings)
class XySaveSettingsAdmin(admin.ModelAdmin):
    list_display = ('user', 'daily_interest_notifications', 'goal_reminders', 'auto_save_notifications', 'preferred_interest_payout')
    list_filter = ('daily_interest_notifications', 'goal_reminders', 'auto_save_notifications', 'preferred_interest_payout')
    search_fields = ('user__username', 'user__email')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('User Information', {
            'fields': ('user',)
        }),
        ('Notification Settings', {
            'fields': ('daily_interest_notifications', 'goal_reminders', 'auto_save_notifications', 'investment_updates')
        }),
        ('Interest Settings', {
            'fields': ('preferred_interest_payout',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['enable_all_notifications', 'disable_all_notifications']
    
    def enable_all_notifications(self, request, queryset):
        updated = queryset.update(
            daily_interest_notifications=True,
            goal_reminders=True,
            auto_save_notifications=True,
            investment_updates=True
        )
        self.message_user(request, f'Enabled notifications for {updated} settings.')
    enable_all_notifications.short_description = 'Enable all notifications'
    
    def disable_all_notifications(self, request, queryset):
        updated = queryset.update(
            daily_interest_notifications=False,
            goal_reminders=False,
            auto_save_notifications=False,
            investment_updates=False
        )
        self.message_user(request, f'Disabled notifications for {updated} settings.')


# Spend and Save Admin Classes
class SpendAndSaveAccountAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'account_number', 'balance', 'is_active', 'savings_percentage',
        'total_interest_earned', 'total_saved_from_spending', 'total_transactions_processed',
        'created_at'
    )
    list_filter = ('is_active', 'savings_percentage', 'created_at')
    search_fields = ('user__username', 'user__email', 'account_number')
    readonly_fields = ('created_at', 'updated_at', 'last_interest_calculation', 'interest_breakdown_info', 'tiered_rates_info')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Account Information', {
            'fields': ('user', 'account_number')
        }),
        ('Financial Information', {
            'fields': ('balance', 'total_interest_earned', 'total_saved_from_spending')
        }),
        ('Spend and Save Configuration', {
            'fields': ('is_active', 'savings_percentage', 'min_transaction_amount')
        }),
        ('Tracking Information', {
            'fields': ('total_transactions_processed', 'last_auto_save_date')
        }),
        ('Withdrawal Settings', {
            'fields': ('default_withdrawal_destination',)
        }),
        ('Interest Information', {
            'fields': ('interest_breakdown_info', 'tiered_rates_info'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def interest_breakdown_info(self, obj):
        """Display current interest breakdown"""
        if obj.balance.amount <= 0:
            return format_html('<span style="color: #6c757d;">No interest (zero balance)</span>')
        
        breakdown = obj.get_interest_breakdown()
        
        html = '<div style="background: #f8f9fa; padding: 10px; border-radius: 5px;">'
        html += '<h4>Current Interest Breakdown:</h4>'
        
        for tier_name, tier_data in breakdown.items():
            if tier_name != 'total_interest':
                html += f'<p><strong>{tier_name.replace("_", " ").title()}:</strong> '
                html += f'₦{tier_data["amount"]:,.2f} at {tier_data["rate"]}% p.a. = ₦{tier_data["interest"]:,.2f}</p>'
        
        html += f'<hr><p><strong>Total Daily Interest:</strong> ₦{breakdown["total_interest"]:,.2f}</p>'
        html += '</div>'
        
        return mark_safe(html)
    interest_breakdown_info.short_description = 'Interest Breakdown'
    
    def tiered_rates_info(self, obj):
        """Display tiered rates information"""
        html = '<div style="background: #e9ecef; padding: 10px; border-radius: 5px;">'
        html += '<h4>Tiered Interest Rates:</h4>'
        html += '<p><strong>Tier 1:</strong> First ₦10,000 at 20% p.a</p>'
        html += '<p><strong>Tier 2:</strong> ₦10,001 - ₦100,000 at 16% p.a</p>'
        html += '<p><strong>Tier 3:</strong> Above ₦100,000 at 8% p.a</p>'
        html += '</div>'
        
        return mark_safe(html)
    tiered_rates_info.short_description = 'Tiered Rates Info'
    
    actions = ['activate_accounts', 'deactivate_accounts', 'calculate_interest', 'process_daily_interest']
    
    def activate_accounts(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'Activated {updated} Spend and Save accounts.')
    activate_accounts.short_description = 'Activate selected accounts'
    
    def deactivate_accounts(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'Deactivated {updated} Spend and Save accounts.')
    deactivate_accounts.short_description = 'Deactivate selected accounts'
    
    def calculate_interest(self, request, queryset):
        from .spend_and_save_services import SpendAndSaveService
        
        processed = 0
        for account in queryset:
            try:
                interest_tx = SpendAndSaveService.calculate_and_credit_interest(account.user)
                if interest_tx:
                    processed += 1
            except Exception as e:
                self.message_user(request, f'Error processing interest for {account.user.username}: {str(e)}', level='ERROR')
        
        self.message_user(request, f'Calculated and credited interest for {processed} accounts.')
    calculate_interest.short_description = 'Calculate and credit interest'
    
    def process_daily_interest(self, request, queryset):
        from .spend_and_save_services import SpendAndSaveInterestService
        
        processed = SpendAndSaveInterestService.process_daily_interest_payout()
        self.message_user(request, f'Processed daily interest for {processed} accounts.')
    process_daily_interest.short_description = 'Process daily interest for all active accounts'


class SpendAndSaveTransactionAdmin(admin.ModelAdmin):
    list_display = (
        'reference', 'spend_and_save_account', 'transaction_type', 'amount',
        'balance_before', 'balance_after', 'savings_percentage_applied', 'created_at'
    )
    list_filter = ('transaction_type', 'created_at', 'withdrawal_destination')
    search_fields = ('reference', 'spend_and_save_account__user__username', 'description')
    readonly_fields = ('created_at', 'interest_breakdown_display')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Transaction Information', {
            'fields': ('reference', 'spend_and_save_account', 'transaction_type', 'amount')
        }),
        ('Balance Information', {
            'fields': ('balance_before', 'balance_after')
        }),
        ('Auto-Save Details', {
            'fields': ('original_transaction_id', 'original_transaction_amount', 'savings_percentage_applied'),
            'classes': ('collapse',)
        }),
        ('Withdrawal Details', {
            'fields': ('withdrawal_destination', 'destination_account'),
            'classes': ('collapse',)
        }),
        ('Interest Information', {
            'fields': ('interest_earned', 'interest_breakdown_display'),
            'classes': ('collapse',)
        }),
        ('Description', {
            'fields': ('description',)
        }),
        ('Timestamps', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    
    def interest_breakdown_display(self, obj):
        """Display interest breakdown in a readable format"""
        if not obj.interest_breakdown:
            return "No interest breakdown available"
        
        html = '<div style="background: #f8f9fa; padding: 10px; border-radius: 5px;">'
        html += '<h4>Interest Breakdown:</h4>'
        
        for tier_name, tier_data in obj.interest_breakdown.items():
            if tier_name != 'total_interest':
                html += f'<p><strong>{tier_name.replace("_", " ").title()}:</strong> '
                html += f'₦{tier_data.get("amount", 0):,.2f} at {tier_data.get("rate", 0)}% p.a. = ₦{tier_data.get("interest", 0):,.2f}</p>'
        
        html += f'<hr><p><strong>Total Interest:</strong> ₦{obj.interest_breakdown.get("total_interest", 0):,.2f}</p>'
        html += '</div>'
        
        return mark_safe(html)
    interest_breakdown_display.short_description = 'Interest Breakdown'


class SpendAndSaveSettingsAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'auto_save_notifications', 'interest_notifications', 'withdrawal_notifications',
        'preferred_savings_percentage', 'default_withdrawal_destination', 'interest_payout_frequency'
    )
    list_filter = (
        'auto_save_notifications', 'interest_notifications', 'withdrawal_notifications',
        'default_withdrawal_destination', 'interest_payout_frequency'
    )
    search_fields = ('user__username', 'user__email')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('User Information', {
            'fields': ('user',)
        }),
        ('Notification Settings', {
            'fields': ('auto_save_notifications', 'interest_notifications', 'withdrawal_notifications')
        }),
        ('Auto-Save Preferences', {
            'fields': ('preferred_savings_percentage', 'min_transaction_threshold')
        }),
        ('Withdrawal Preferences', {
            'fields': ('default_withdrawal_destination',)
        }),
        ('Interest Preferences', {
            'fields': ('interest_payout_frequency',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['enable_all_notifications', 'disable_all_notifications']
    
    def enable_all_notifications(self, request, queryset):
        updated = queryset.update(
            auto_save_notifications=True,
            interest_notifications=True,
            withdrawal_notifications=True
        )
        self.message_user(request, f'Enabled notifications for {updated} settings.')
    enable_all_notifications.short_description = 'Enable all notifications'
    
    def disable_all_notifications(self, request, queryset):
        updated = queryset.update(
            auto_save_notifications=False,
            interest_notifications=False,
            withdrawal_notifications=False
        )
        self.message_user(request, f'Disabled notifications for {updated} settings.')
    disable_all_notifications.short_description = 'Disable all notifications'


# Register Spend and Save models
admin.site.register(SpendAndSaveAccount, SpendAndSaveAccountAdmin)
admin.site.register(SpendAndSaveTransaction, SpendAndSaveTransactionAdmin)
admin.site.register(SpendAndSaveSettings, SpendAndSaveSettingsAdmin)


# Target Saving Admin Classes

class TargetSavingDepositInline(admin.TabularInline):
    """Inline admin for target saving deposits"""
    model = TargetSavingDeposit
    extra = 0
    readonly_fields = ('deposit_date',)
    fields = ('amount', 'notes', 'deposit_date')
    can_delete = False


class TargetSavingWithdrawalInline(admin.TabularInline):
    """Inline admin for target saving withdrawals"""
    model = TargetSavingWithdrawal
    extra = 0
    readonly_fields = ('withdrawal_date',)
    fields = ('amount', 'notes', 'withdrawal_date')
    can_delete = False


@admin.register(TargetSaving)
class TargetSavingAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'user', 'account_number', 'source', 'category', 'target_amount', 'current_amount', 
        'progress_percentage', 'frequency', 'is_active', 'is_completed', 
        'days_remaining', 'is_overdue', 'created_at'
    )
    list_filter = (
        'category', 'source', 'frequency', 'is_active', 'is_completed', 'created_at',
        'preferred_deposit_day'
    )
    search_fields = ('name', 'account_number', 'user__username', 'user__email')
    readonly_fields = (
        'created_at', 'updated_at', 'progress_percentage', 'remaining_amount',
        'days_remaining', 'is_overdue', 'daily_target', 'weekly_target', 
        'monthly_target', 'total_deposits_count'
    )
    ordering = ('-created_at',)
    inlines = [TargetSavingDepositInline, TargetSavingWithdrawalInline]
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('user', 'name', 'category', 'account_number', 'source')
        }),
        ('Target Details', {
            'fields': ('target_amount', 'frequency', 'preferred_deposit_day')
        }),
        ('Time Period', {
            'fields': ('start_date', 'end_date')
        }),
        ('Progress Information', {
            'fields': ('current_amount', 'progress_percentage', 'remaining_amount', 'total_deposits_count')
        }),
        ('Calculated Targets', {
            'fields': ('daily_target', 'weekly_target', 'monthly_target'),
            'classes': ('collapse',)
        }),
        ('Status Information', {
            'fields': ('is_active', 'is_completed', 'days_remaining', 'is_overdue')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def progress_percentage(self, obj):
        """Display progress percentage with color coding"""
        percentage = obj.progress_percentage
        if percentage >= 100:
            color = '#28a745'  # Green
        elif percentage >= 75:
            color = '#17a2b8'  # Blue
        elif percentage >= 50:
            color = '#ffc107'  # Yellow
        else:
            color = '#dc3545'  # Red
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}%</span>',
            color, f"{percentage:.1f}"
        )
    progress_percentage.short_description = 'Progress'
    
    def days_remaining(self, obj):
        """Display days remaining with color coding"""
        days = obj.days_remaining
        if days < 0:
            color = '#dc3545'  # Red (overdue)
        elif days <= 7:
            color = '#ffc107'  # Yellow (urgent)
        else:
            color = '#28a745'  # Green (good)
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} days</span>',
            color, days
        )
    days_remaining.short_description = 'Days Left'
    
    def is_overdue(self, obj):
        """Display overdue status"""
        return obj.is_overdue
    is_overdue.boolean = True
    is_overdue.short_description = 'Overdue'
    
    def total_deposits_count(self, obj):
        """Display total deposits count"""
        return obj.deposits.count()
    total_deposits_count.short_description = 'Total Deposits'
    
    def daily_target(self, obj):
        """Display daily target amount"""
        try:
            target = obj.daily_target
            if target == 0:
                return format_html('<span style="color: #6c757d;">₦0.00</span>')
            return format_html('<span style="color: #17a2b8; font-weight: bold;">₦{}</span>', f"{target:,.2f}")
        except (TypeError, AttributeError):
            return format_html('<span style="color: #6c757d;">₦0.00</span>')
    daily_target.short_description = 'Daily Target'
    
    def weekly_target(self, obj):
        """Display weekly target amount"""
        try:
            target = obj.weekly_target
            if target == 0:
                return format_html('<span style="color: #6c757d;">₦0.00</span>')
            return format_html('<span style="color: #17a2b8; font-weight: bold;">₦{}</span>', f"{target:,.2f}")
        except (TypeError, AttributeError):
            return format_html('<span style="color: #6c757d;">₦0.00</span>')
    weekly_target.short_description = 'Weekly Target'
    
    def monthly_target(self, obj):
        """Display monthly target amount"""
        try:
            target = obj.monthly_target
            if target == 0:
                return format_html('<span style="color: #6c757d;">₦0.00</span>')
            return format_html('<span style="color: #17a2b8; font-weight: bold;">₦{}</span>', f"{target:,.2f}")
        except (TypeError, AttributeError):
            return format_html('<span style="color: #6c757d;">₦0.00</span>')
    monthly_target.short_description = 'Monthly Target'
    
    actions = [
        'activate_targets', 'deactivate_targets', 'mark_as_completed', 
        'export_target_data', 'send_reminder_notifications'
    ]
    
    def activate_targets(self, request, queryset):
        """Activate selected targets"""
        updated = queryset.update(is_active=True)
        self.message_user(request, f'Activated {updated} target savings.')
    activate_targets.short_description = 'Activate selected targets'
    
    def deactivate_targets(self, request, queryset):
        """Deactivate selected targets"""
        updated = queryset.update(is_active=False)
        self.message_user(request, f'Deactivated {updated} target savings.')
    deactivate_targets.short_description = 'Deactivate selected targets'
    
    def mark_as_completed(self, request, queryset):
        """Mark targets as completed"""
        updated = queryset.update(is_completed=True, is_active=False)
        self.message_user(request, f'Marked {updated} target savings as completed.')
    mark_as_completed.short_description = 'Mark as completed'
    
    def export_target_data(self, request, queryset):
        """Export target saving data to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="target_savings.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Name', 'User', 'Category', 'Target Amount', 'Current Amount', 
            'Progress %', 'Frequency', 'Start Date', 'End Date', 'Days Remaining',
            'Is Active', 'Is Completed', 'Created At'
        ])
        
        for target in queryset:
            writer.writerow([
                target.name,
                target.user.username,
                target.get_category_display(),
                target.target_amount,
                target.current_amount,
                f"{target.progress_percentage:.1f}%",
                target.get_frequency_display(),
                target.start_date,
                target.end_date,
                target.days_remaining,
                'Yes' if target.is_active else 'No',
                'Yes' if target.is_completed else 'No',
                target.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    export_target_data.short_description = 'Export to CSV'
    
    def send_reminder_notifications(self, request, queryset):
        """Send reminder notifications for selected targets"""
        from .target_saving_services import TargetSavingNotificationService
        
        sent_count = 0
        for target in queryset:
            if target.is_active and not target.is_completed:
                try:
                    TargetSavingNotificationService.send_target_reminder_notification(
                        target.user, target, 'admin_reminder'
                    )
                    sent_count += 1
                except Exception as e:
                    self.message_user(
                        request, 
                        f'Error sending reminder for {target.name}: {str(e)}', 
                        level=messages.ERROR
                    )
        
        self.message_user(request, f'Sent reminder notifications for {sent_count} targets.')
    send_reminder_notifications.short_description = 'Send reminder notifications'


@admin.register(TargetSavingDeposit)
class TargetSavingDepositAdmin(admin.ModelAdmin):
    list_display = (
        'target_saving', 'amount', 'deposit_date', 'notes', 'user_info'
    )
    list_filter = ('deposit_date', 'target_saving__category', 'target_saving__frequency')
    search_fields = (
        'target_saving__name', 'target_saving__user__username', 
        'target_saving__user__email', 'notes'
    )
    readonly_fields = ('deposit_date', 'target_saving_link')
    ordering = ('-deposit_date',)
    
    fieldsets = (
        ('Deposit Information', {
            'fields': ('target_saving', 'amount', 'notes')
        }),
        ('Target Information', {
            'fields': ('target_saving_link',),
            'classes': ('collapse',)
        }),
        ('Timestamp', {
            'fields': ('deposit_date',),
            'classes': ('collapse',)
        }),
    )
    
    def user_info(self, obj):
        """Display user information"""
        user = obj.target_saving.user
        return f"{user.username} ({user.email})"
    user_info.short_description = 'User'
    
    def target_saving_link(self, obj):
        """Display target saving with link"""
        target = obj.target_saving
        return format_html(
            '<a href="{}">{}</a> - {}% complete',
            reverse('admin:bank_targetsaving_change', args=[target.id]),
            target.name,
            f"{target.progress_percentage:.1f}"
        )
    target_saving_link.short_description = 'Target Saving'
    
    actions = ['export_deposits_csv', 'export_deposits_xlsx']
    
    def export_deposits_csv(self, request, queryset):
        """Export deposits to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="target_saving_deposits.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Target Name', 'User', 'Category', 'Amount', 'Notes', 'Deposit Date'
        ])
        
        for deposit in queryset:
            target = deposit.target_saving
            writer.writerow([
                target.name,
                target.user.username,
                target.get_category_display(),
                deposit.amount,
                deposit.notes or '',
                deposit.deposit_date.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    export_deposits_csv.short_description = 'Export to CSV'
    
    def export_deposits_xlsx(self, request, queryset):
        """Export deposits to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Target Saving Deposits'
        
        headers = ['Target Name', 'User', 'Category', 'Amount', 'Notes', 'Deposit Date']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for deposit in queryset:
            target = deposit.target_saving
            ws.append([
                target.name,
                target.user.username,
                target.get_category_display(),
                deposit.amount,
                deposit.notes or '',
                deposit.deposit_date.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=target_saving_deposits.xlsx'
        wb.save(response)
        return response
    export_deposits_xlsx.short_description = 'Export to Excel'


@admin.register(TargetSavingWithdrawal)
class TargetSavingWithdrawalAdmin(admin.ModelAdmin):
    list_display = (
        'target_saving', 'amount', 'withdrawal_date', 'notes', 'user_info'
    )
    list_filter = ('withdrawal_date', 'target_saving__category', 'target_saving__frequency')
    search_fields = (
        'target_saving__name', 'target_saving__user__username',
        'target_saving__user__email', 'notes'
    )
    readonly_fields = ('withdrawal_date',)
    ordering = ('-withdrawal_date',)

    def user_info(self, obj):
        """Display user information"""
        user = obj.target_saving.user
        return f"{user.username} ({user.email})"
    user_info.short_description = 'User'


# Note: TargetSavingCategory and TargetSavingFrequency are TextChoices classes, not Django models
# They cannot be registered with Django admin as they don't have _meta attributes
# These choices are available in the TargetSaving model's category and frequency fields

# TargetSaving and TargetSavingDeposit are registered via @admin.register decorators above


# Update the admin URL configuration to include the staff dashboard
def get_admin_urls():
    from django.urls import path
    return [
        path('dashboard/', admin_dashboard, name='admin-dashboard'),
        path('staff-dashboard/', staff_dashboard, name='staff-dashboard'),
        path('interest-calculator/', interest_calculator_admin, name='interest-calculator'),
    ]

# Store the original get_urls method
original_get_urls = admin.site.get_urls

# Override with our custom function
def custom_get_urls():
    return get_admin_urls() + original_get_urls()

admin.site.get_urls = custom_get_urls

# Fixed Savings Admin Classes
class FixedSavingsTransactionInline(admin.TabularInline):
    """Inline admin for fixed savings transactions"""
    model = FixedSavingsTransaction
    extra = 0
    readonly_fields = ('created_at', 'balance_before', 'balance_after', 'reference')
    fields = ('transaction_type', 'amount', 'interest_earned', 'description', 'created_at')
    can_delete = False

@admin.register(FixedSavingsAccount)
class FixedSavingsAccountAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'account_number', 'amount', 'source', 'purpose', 'interest_rate',
        'maturity_amount', 'days_remaining', 'is_active', 'is_matured', 'is_paid_out',
        'auto_renewal_enabled', 'created_at'
    )
    list_filter = (
        'source', 'purpose', 'is_active', 'is_matured', 'is_paid_out', 
        'auto_renewal_enabled', 'created_at'
    )
    search_fields = ('user__username', 'user__email', 'account_number', 'purpose_description')
    readonly_fields = (
        'created_at', 'updated_at', 'matured_at', 'paid_out_at', 'duration_days',
        'days_remaining', 'is_mature', 'can_be_paid_out', 'interest_rate', 'maturity_amount'
    )
    ordering = ('-created_at',)
    inlines = [FixedSavingsTransactionInline]
    
    fieldsets = (
        ('Account Information', {
            'fields': ('user', 'account_number')
        }),
        ('Fixed Savings Details', {
            'fields': ('amount', 'source', 'purpose', 'purpose_description')
        }),
        ('Time Period', {
            'fields': ('start_date', 'payback_date', 'duration_days')
        }),
        ('Interest Information', {
            'fields': ('interest_rate', 'total_interest_earned', 'maturity_amount')
        }),
        ('Status Information', {
            'fields': ('is_active', 'is_matured', 'is_paid_out', 'days_remaining', 'is_mature', 'can_be_paid_out')
        }),
        ('Auto-Renewal', {
            'fields': ('auto_renewal_enabled',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at', 'matured_at', 'paid_out_at'),
            'classes': ('collapse',)
        }),
    )
    
    def duration_days(self, obj):
        """Display duration in days"""
        return obj.duration_days
    duration_days.short_description = 'Duration (Days)'
    
    def days_remaining(self, obj):
        """Display days remaining with color coding"""
        days = obj.days_remaining
        if days < 0:
            color = '#dc3545'  # Red (overdue)
        elif days <= 7:
            color = '#ffc107'  # Yellow (urgent)
        else:
            color = '#28a745'  # Green (good)
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} days</span>',
            color, days
        )
    days_remaining.short_description = 'Days Left'
    
    def is_mature(self, obj):
        """Display maturity status"""
        return obj.is_mature
    is_mature.boolean = True
    is_mature.short_description = 'Mature'
    
    def can_be_paid_out(self, obj):
        """Display payout eligibility"""
        return obj.can_be_paid_out
    can_be_paid_out.boolean = True
    can_be_paid_out.short_description = 'Can Pay Out'
    
    actions = [
        'mark_as_matured', 'pay_out_fixed_savings', 'enable_auto_renewal', 
        'disable_auto_renewal', 'export_fixed_savings_data'
    ]
    
    def mark_as_matured(self, request, queryset):
        """Mark selected fixed savings as matured"""
        matured_count = 0
        for fixed_savings in queryset:
            if fixed_savings.mark_as_matured():
                matured_count += 1
        
        self.message_user(
            request, 
            f'Successfully marked {matured_count} fixed savings as matured.'
        )
    mark_as_matured.short_description = 'Mark as Matured'
    
    def pay_out_fixed_savings(self, request, queryset):
        """Pay out matured fixed savings"""
        paid_out_count = 0
        for fixed_savings in queryset:
            if fixed_savings.pay_out():
                paid_out_count += 1
        
        self.message_user(
            request, 
            f'Successfully paid out {paid_out_count} fixed savings.'
        )
    pay_out_fixed_savings.short_description = 'Pay Out Fixed Savings'
    
    def enable_auto_renewal(self, request, queryset):
        """Enable auto-renewal for selected fixed savings"""
        updated = queryset.update(auto_renewal_enabled=True)
        self.message_user(
            request, 
            f'Successfully enabled auto-renewal for {updated} fixed savings.'
        )
    enable_auto_renewal.short_description = 'Enable Auto-Renewal'
    
    def disable_auto_renewal(self, request, queryset):
        """Disable auto-renewal for selected fixed savings"""
        updated = queryset.update(auto_renewal_enabled=False)
        self.message_user(
            request, 
            f'Successfully disabled auto-renewal for {updated} fixed savings.'
        )
    disable_auto_renewal.short_description = 'Disable Auto-Renewal'
    
    def export_fixed_savings_data(self, request, queryset):
        """Export fixed savings data to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fixed_savings_data.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'User', 'Account Number', 'Amount', 'Source', 'Purpose', 'Interest Rate',
            'Maturity Amount', 'Start Date', 'Payback Date', 'Duration Days',
            'Days Remaining', 'Is Active', 'Is Matured', 'Is Paid Out',
            'Auto Renewal', 'Created At'
        ])
        
        for fixed_savings in queryset:
            writer.writerow([
                fixed_savings.user.username,
                fixed_savings.account_number,
                fixed_savings.amount,
                fixed_savings.get_source_display(),
                fixed_savings.get_purpose_display(),
                f"{fixed_savings.interest_rate}%",
                fixed_savings.maturity_amount,
                fixed_savings.start_date,
                fixed_savings.payback_date,
                fixed_savings.duration_days,
                fixed_savings.days_remaining,
                fixed_savings.is_active,
                fixed_savings.is_matured,
                fixed_savings.is_paid_out,
                fixed_savings.auto_renewal_enabled,
                fixed_savings.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    export_fixed_savings_data.short_description = 'Export to CSV'

@admin.register(FixedSavingsTransaction)
class FixedSavingsTransactionAdmin(admin.ModelAdmin):
    list_display = (
        'reference', 'fixed_savings_account', 'transaction_type', 'amount',
        'interest_earned', 'source_account', 'created_at'
    )
    list_filter = ('transaction_type', 'source_account', 'created_at')
    search_fields = (
        'reference', 'fixed_savings_account__user__username', 
        'fixed_savings_account__account_number', 'description'
    )
    readonly_fields = ('created_at', 'balance_before', 'balance_after', 'reference')
    ordering = ('-created_at',)
    
    fieldsets = (
        ('Transaction Information', {
            'fields': ('fixed_savings_account', 'transaction_type', 'amount', 'description')
        }),
        ('Balance Information', {
            'fields': ('balance_before', 'balance_after', 'reference')
        }),
        ('Interest Information', {
            'fields': ('interest_earned', 'interest_rate_applied')
        }),
        ('Source Information', {
            'fields': ('source_account', 'source_transaction_id'),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('metadata',),
            'classes': ('collapse',)
        }),
        ('Timestamp', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['export_transactions_csv', 'export_transactions_xlsx']
    
    def export_transactions_csv(self, request, queryset):
        """Export transactions to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fixed_savings_transactions.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Reference', 'User', 'Account Number', 'Transaction Type', 'Amount',
            'Interest Earned', 'Source Account', 'Description', 'Created At'
        ])
        
        for transaction in queryset:
            writer.writerow([
                transaction.reference,
                transaction.fixed_savings_account.user.username,
                transaction.fixed_savings_account.account_number,
                transaction.get_transaction_type_display(),
                transaction.amount,
                transaction.interest_earned,
                transaction.get_source_account_display() if transaction.source_account else '',
                transaction.description or '',
                transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    export_transactions_csv.short_description = 'Export to CSV'
    
    def export_transactions_xlsx(self, request, queryset):
        """Export transactions to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fixed Savings Transactions'
        
        headers = [
            'Reference', 'User', 'Account Number', 'Transaction Type', 'Amount',
            'Interest Earned', 'Source Account', 'Description', 'Created At'
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for transaction in queryset:
            ws.append([
                transaction.reference,
                transaction.fixed_savings_account.user.username,
                transaction.fixed_savings_account.account_number,
                transaction.get_transaction_type_display(),
                transaction.amount,
                transaction.interest_earned,
                transaction.get_source_account_display() if transaction.source_account else '',
                transaction.description or '',
                transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=fixed_savings_transactions.xlsx'
        wb.save(response)
        return response
    export_transactions_xlsx.short_description = 'Export to Excel'

@admin.register(FixedSavingsSettings)
class FixedSavingsSettingsAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'maturity_notifications', 'interest_notifications', 
        'auto_renewal_notifications', 'default_auto_renewal', 
        'default_renewal_duration', 'default_source'
    )
    list_filter = (
        'maturity_notifications', 'interest_notifications', 'auto_renewal_notifications',
        'default_auto_renewal', 'default_source'
    )
    search_fields = ('user__username', 'user__email')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('User Information', {
            'fields': ('user',)
        }),
        ('Notification Settings', {
            'fields': ('maturity_notifications', 'interest_notifications', 'auto_renewal_notifications')
        }),
        ('Auto-Renewal Preferences', {
            'fields': ('default_auto_renewal', 'default_renewal_duration')
        }),
        ('Source Preferences', {
            'fields': ('default_source',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['enable_all_notifications', 'disable_all_notifications']
    
    def enable_all_notifications(self, request, queryset):
        """Enable all notifications for selected users"""
        updated = queryset.update(
            maturity_notifications=True,
            interest_notifications=True,
            auto_renewal_notifications=True
        )
        self.message_user(
            request, 
            f'Successfully enabled all notifications for {updated} users.'
        )
    enable_all_notifications.short_description = 'Enable All Notifications'
    
    def disable_all_notifications(self, request, queryset):
        """Disable all notifications for selected users"""
        updated = queryset.update(
            maturity_notifications=False,
            interest_notifications=False,
            auto_renewal_notifications=False
        )
        self.message_user(
            request, 
            f'Successfully disabled all notifications for {updated} users.'
        )
